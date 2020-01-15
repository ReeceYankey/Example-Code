import openpyxl as xl
from copy import copy
from collections import Counter


class WeightedSheetHandler:
    def __init__(self, ws):
        self.ws = ws

    def is_header(self, row, column):
        try:
            return self.ws.cell(row=row, column=column + 1).value == 'Due Date' and \
                   '=B' in self.ws.cell(row=row, column=column).value
        except AttributeError or ValueError:
            return False

    def is_totals_row(self, row, column):
        try:
            return self.ws.cell(row=row, column=column + 1).value == 'Total:' and \
                   '=B' in self.ws.cell(row=row, column=column).value
        except AttributeError or ValueError:
            return False

    def is_end_of_section(self, row, column):
        return self.ws.cell(row=row, column=column).value is not None and \
               '*If you need to add a row' in self.ws.cell(row=row,column=column).value

    def get_header_rows(self):
        ws = self.ws

        header_rows = []
        for r in range(36, ws.max_row + 1):
            if self.is_header(r, 2):
                header_rows.append(r)
        return header_rows

    def get_totals_rows(self):
        ws = self.ws

        totals_rows = []
        for r in range(37, ws.max_row + 1):
            if self.is_totals_row(r, 2):
                totals_rows.append(r)
        return totals_rows

    def add_row(self, row, count=1):
        ws = self.ws

        ws.insert_rows(row, amount=count)

        # ----update pointers at top (C16:D25)-----
        ending_rows = self.get_totals_rows()
        header_rows = self.get_header_rows()

        for r in range(16, 26):
            for c in range(3, 5):
                cell = ws.cell(row=r, column=c)
                ending_row_index = (r - 16) // 3
                cell.value = cell.value[:2] + str(ending_rows[ending_row_index])

        data_row_starts = []  # the starts for the areas that hold assignment data
        data_row_ends = []  # the ends for the areas that hold assignment data
        for i in header_rows:
            data_row_starts.append(i + 1)
        for i in ending_rows:
            data_row_ends.append(i - 3)

        for i in range(len(ending_rows)):
            r = ending_rows[i]
            for c in [4, 10, 17]:
                cell = ws.cell(row=r, column=c)
                column_letter = xl.utils.get_column_letter(c)
                cell.value = '=SUM({}:{})'.format(column_letter + str(data_row_starts[i]),
                                                  column_letter + str(data_row_ends[i]))
            for c in [5, 11, 18]:
                cell = ws.cell(row=r, column=c)
                column_letter_1 = xl.utils.get_column_letter(c - 1)
                column_letter_2 = xl.utils.get_column_letter(c)
                cell.value = '=SUMIF({}:{},">=0",{}:{})'.format(column_letter_1 + str(data_row_starts[i]),
                                                                column_letter_1 + str(data_row_ends[i]),
                                                                column_letter_2 + str(data_row_starts[i]),
                                                                column_letter_2 + str(data_row_ends[i]))
        # removes unnecessary pointers from above code
        ws['J{}'.format(ending_rows[-1])].value = None
        ws['K{}'.format(ending_rows[-1])].value = None
        ws['Q{}'.format(ending_rows[-1])].value = None
        ws['R{}'.format(ending_rows[-1])].value = None

        # -----style cells-----
        for i in [2, 8, 15]:
            for col in range(i, i + 4):
                ws.cell(row=row, column=col)._style = copy(ws.cell(row=row - 1, column=col)._style)

    def unmerge_ending_cells(self):
        ws = self.ws

        totals_rows = self.get_totals_rows()
        for c in [2, 8, 15]:
            for r in totals_rows:
                try:
                    ws.unmerge_cells(start_row=r-2, start_column=c, end_row=r - 1, end_column=c + 3)
                except ValueError as e:
                    print(e)

    def merge_ending_cells(self):
        ws = self.ws

        totals_rows = self.get_totals_rows()
        for c in [2, 8, 15]:
            for r in totals_rows:
                ws.merge_cells(start_row=r-2, start_column=c, end_row=r - 1, end_column=c + 3)

    def update(self, table):
        ws = self.ws

        # TODO: fix potential conflict: table can be either list or pandas.core.series.Series

        self.unmerge_ending_cells()  #merged cells cannot be modified, so all cells that are shifted must be unmerged then remerged at the end
        # set categories and weightings
        categories = list(Counter(table['type']))
        for r in range(len(categories)):
            self.ws.cell(row=r + 16, column=2).value = categories[r]

        # add assignments to corresponding sections
        for sheet_col in [2, 8, 15]:
            sheet_row = 36
            while sheet_row < ws.max_row:
                sheet_row += 1  # loop begins at 37
                cell = ws.cell(row=sheet_row, column=sheet_col)

                if not self.is_header(sheet_row, sheet_col):
                    continue

                category_title = ws[cell.value[1:]].value  # gets the value of the cell that the cell points to
                if category_title is None:
                    continue
                # t = table['type'].tolist()
                # assert category_title in table['type'].tolist()

                # add data to section
                sheet_row += 1
                for i in range(len(table['type'])):
                    if table['type'][i] == category_title:
                        if self.is_end_of_section(sheet_row, sheet_col):
                            self.add_row(sheet_row)
                        ws.cell(row=sheet_row, column=sheet_col).value = table['name'][i]
                        ws.cell(row=sheet_row, column=sheet_col + 1).value = table['date'][i]
                        ws.cell(row=sheet_row, column=sheet_col + 2).value = table['score'][i]
                        ws.cell(row=sheet_row, column=sheet_col + 3).value = table['max_score'][i]
                        sheet_row += 1

                # clear old data from section
                # print(type(ws.cell(row=sheet_row, column=sheet_col).value))
                while not self.is_end_of_section(sheet_row, sheet_col):
                    ws.cell(row=sheet_row, column=sheet_col).value = None
                    ws.cell(row=sheet_row, column=sheet_col + 1).value = None
                    ws.cell(row=sheet_row, column=sheet_col + 2).value = None
                    ws.cell(row=sheet_row, column=sheet_col + 3).value = None
                    sheet_row += 1

        self.merge_ending_cells()


class PointSheetHandler:
    def __init__(self, ws):
        self.ws = ws

    def is_totals_row(self, row):
        return self.ws.cell(row=row, column=2).value == 'Total'

    def get_totals_row(self):
        ws = self.ws

        for r in range(37, ws.max_row + 1):
            if self.is_totals_row(r):
                return r
    
    def add_row(self, row, count=1):
        ws = self.ws

        ws.insert_rows(row, amount=count)

        # ----update pointers at top-----
        totals_row = self.get_totals_row()

        cell = ws['K8']
        cell.value = '={} / {} * 100'.format('D' + str(totals_row),
                                             'G' + str(totals_row))
        cell = ws['K9']
        cell.value = '={}'.format('D' + str(totals_row))

        data_row_start = 16  # the start for the area that holds assignment data
        data_row_end = totals_row - 1  # the end for the area that holds assignment data

        cell = ws.cell(row=totals_row, column=4)
        cell.value = '=SUM({}:{})'.format('D' + str(data_row_start),
                                          'D' + str(data_row_end))
        cell = ws.cell(row=totals_row, column=7)
        cell.value = '=SUMIF({}:{},">=0",{}:{})'.format('D' + str(data_row_start),
                                                        'D' + str(data_row_end),
                                                        'G' + str(data_row_start),
                                                        'G' + str(data_row_end))
        cell = ws['R16']
        cell.value = '=SUM({}:{})'.format('G' + str(data_row_start),
                                          'G' + str(data_row_end))

        # -----style cells-----
        for col in range(1, 8):
            ws.cell(row=row, column=col)._style = copy(ws.cell(row=row - 1, column=col)._style)

    def unmerge_ending_cells(self):
        ws = self.ws

        totals_row = self.get_totals_row()
        ws.unmerge_cells(start_row=totals_row+1, start_column=1, end_row=totals_row+2, end_column=7)

    def merge_ending_cells(self):
        ws = self.ws

        totals_row = self.get_totals_row()
        print('merging', totals_row+1, 1, totals_row+2, 7)
        ws.merge_cells(start_row=totals_row+1, start_column=1, end_row=totals_row+2, end_column=7)

    def update(self, table):
        ws = self.ws
        
        self.unmerge_ending_cells()
        
        # add data to section
        sheet_row = 16
        for i in range(len(table['name'])):
            if self.is_totals_row(sheet_row):
                self.add_row(sheet_row)
            ws.cell(row=sheet_row, column=2).value = table['name'][i]
            ws.cell(row=sheet_row, column=3).value = table['date'][i]
            ws.cell(row=sheet_row, column=4).value = table['score'][i]
            ws.cell(row=sheet_row, column=7).value = table['max_score'][i]
            sheet_row += 1

        # clear old data from section
        # print(type(ws.cell(row=sheet_row, column=sheet_col).value))
        while not self.is_totals_row(sheet_row):
            ws.cell(row=sheet_row, column=2).value = None
            ws.cell(row=sheet_row, column=3).value = None
            ws.cell(row=sheet_row, column=4).value = None
            ws.cell(row=sheet_row, column=7).value = None
            sheet_row += 1

        self.merge_ending_cells()
